import os
import re
import sys
import pandas as pd
import joblib
import streamlit as st
from openpyxl import load_workbook

# === Functions ===
def collect_data(base_folder):
    all_data = pd.DataFrame()
    total_files = 0
    collected_files = []

    for day in range(1, 32):
        day_folder = os.path.join(base_folder, f"{day:02d}")
        if not os.path.isdir(day_folder):
            continue

        excel_files = [
            f for f in os.listdir(day_folder)
            if "VOLTAGE" in f.upper() and f.endswith((".xlsx", ".xlsm"))
        ]

        total_files += len(excel_files)

        for file_name in excel_files:
            file_path = os.path.join(day_folder, file_name)
            try:
                workbook = load_workbook(file_path, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    date_data = [
                        tuple(cell.value for cell in row)
                        for row in sheet.iter_rows(min_row=13, max_row=32, min_col=4, max_col=5)
                    ]

                    cable_data = [
                        tuple(cell.value for cell in row)
                        for row in sheet.iter_rows(min_row=13, max_row=32, min_col=12, max_col=18)
                    ]

                    for i in range(len(date_data)):
                        date_row = date_data[i]
                        cable_row = cable_data[i]

                        if any(date_row) and any(cable_row):
                            for j in range(len(cable_row)):
                                if cable_row[j]:
                                    all_data = pd.concat([all_data, pd.DataFrame([{
                                        'File': file_name,
                                        'Sheet': sheet_name,
                                        'Date1': date_row[0],
                                        'Date2': date_row[1],
                                        'Cable Name': cable_row[j]
                                    }])], ignore_index=True)

                collected_files.append(file_name)
            except Exception as e:
                st.warning(f"Error in {file_name}: {e}")

    if all_data.empty:
        st.error("No data collected. Check folder structure or file names.")
        return None

    return all_data

def preprocess_data(all_data):
    if 'Date1' in all_data.columns:
        all_data['Date1'] = pd.to_datetime(all_data['Date1'], errors='coerce')
        all_data = all_data.dropna(subset=['Date1']).sort_values(by='Date1')
    else:
        st.error("Missing 'Date1' column")
    return all_data

def predict_cable_type(all_data, model_path="model.pkl"):
    model = joblib.load(model_path)
    df_new = all_data.rename(columns={'Date1': 'Date'})
    cable_names_list = df_new['Cable Name'].tolist()

    predicted = model.predict(cable_names_list)
    proba = model.predict_proba(cable_names_list)
    classes = model.classes_

    confidence, top1, top2, top3 = [], [], [], []
    for i, row in enumerate(proba):
        pred_class = predicted[i]
        pred_index = list(classes).index(pred_class)
        confidence.append(round(row[pred_index] * 100, 2))
        top_classes = sorted(zip(classes, row), key=lambda x: x[1], reverse=True)[:3]
        top1.append(f"{top_classes[0][0]} ({round(top_classes[0][1]*100,1)}%)")
        top2.append(f"{top_classes[1][0]} ({round(top_classes[1][1]*100,1)}%)")
        top3.append(f"{top_classes[2][0]} ({round(top_classes[2][1]*100,1)}%)")

    df_new['Predicted Type'] = predicted
    df_new['Confidence (%)'] = confidence
    df_new['Top 1'] = top1
    df_new['Top 2'] = top2
    df_new['Top 3'] = top3

    group_order = [
        "Control & Instrument cable", "60227 IEC XX", "Signal & Power cable",
        "H1Z2Z2", "HXX, HHXX", "Telecom cable", "AVVSS (TOYOTA)",
        "AV, AVS, AVSS", "CO-V, CO-VV", "UL", "MITI cables"
    ]

    df_new['Adjusted Type'] = df_new.apply(
        lambda row: row['Predicted Type'] if row['Confidence (%)'] >= 70 else "อื่นๆ",
        axis=1
    )

    return df_new, group_order

def clean_date_column(df, column_name="Date"):
    df[column_name] = df[column_name].astype(str).str.split("_").str[0]
    df[column_name] = pd.to_datetime(df[column_name], errors='coerce')
    return df

def summarize_data(df_new, group_order):
    daily_summary = df_new.groupby(['Date', 'Adjusted Type']).size().reset_index(name='Count')
    daily_summary_pivot = daily_summary.pivot(index='Date', columns='Adjusted Type', values='Count')
    all_types_with_others = group_order + ["อื่นๆ"]
    for col in all_types_with_others:
        if col not in daily_summary_pivot.columns:
            daily_summary_pivot[col] = pd.NA
    daily_summary_pivot = daily_summary_pivot[all_types_with_others]
    daily_summary_pivot = daily_summary_pivot.sort_index()
    return daily_summary_pivot

def extract_low_confidence(df_new):
    df_others = df_new[df_new['Confidence (%)'] < 70][[
        'Cable Name', 'Date', 'Predicted Type', 'Confidence (%)',
        'Top 1', 'Top 2', 'Top 3', 'File', 'Sheet']]
    df_others = df_others.sort_values(by='Date')
    return df_others

def save_to_excel(df_new, daily_summary_pivot, df_others):
    from io import BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_new.drop(columns=['Adjusted Type']).to_excel(writer, sheet_name='Predicted Types', index=False)
        daily_summary_pivot.to_excel(writer, sheet_name='Daily Summary by Date')
        df_others.to_excel(writer, sheet_name='Others Detail', index=False)
    output.seek(0)
    return output

# === Streamlit Web App ===
st.set_page_config(page_title="Cable Type Predictor", layout="wide")
st.title("🔌 Cable Type Prediction App")

root_path = st.text_input("🔍 Root Path", r"Z:\\QA Center\\Electrical Testing Report")
model_path = st.text_input("🧠 Model Path", r"C:\\Poohri\\Project\\PythonProject\\model.pkl")

if root_path and os.path.isdir(root_path):
    years = sorted([d for d in os.listdir(root_path) if os.path.isdir(os.path.join(root_path, d))])
    selected_year = st.selectbox("📅 เลือกปี", years)

    year_folder = os.path.join(root_path, selected_year)
    months = sorted([d for d in os.listdir(year_folder) if os.path.isdir(os.path.join(year_folder, d))])
    selected_month = st.selectbox("🗓️ เลือกเดือน", months)

    if st.button("▶️ เริ่มประมวลผล"):
        folder_path = os.path.join(root_path, selected_year, selected_month, "#ผลดิบ")
        if os.path.isdir(folder_path):
            with st.spinner("📥 รวบรวมข้อมูล..."):
                all_data = collect_data(folder_path)

            if all_data is not None:
                all_data = preprocess_data(all_data)
                df_new, group_order = predict_cable_type(all_data, model_path)
                df_new = clean_date_column(df_new, "Date")
                daily_summary_pivot = summarize_data(df_new, group_order)
                df_others = extract_low_confidence(df_new)

                st.success("✅ ประมวลผลสำเร็จแล้ว!")
                st.dataframe(daily_summary_pivot)

                excel_file = save_to_excel(df_new, daily_summary_pivot, df_others)
                st.download_button("📁 ดาวน์โหลดผลลัพธ์เป็น Excel", data=excel_file,
                                   file_name="predicted_cable_types.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.error(f"ไม่พบโฟลเดอร์: {folder_path}")
